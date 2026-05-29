"""FactuSOL mapping support for budget-based document organization with Serie support."""

from __future__ import annotations

import re
import unicodedata
from collections import defaultdict
from dataclasses import dataclass, field
from pathlib import Path
from typing import Iterable, Sequence


# OK statuses
OK_STATUSES = {"OK_CLAVEAPP", "OK_ANIO_SERIE_PRESUPUESTO", "OK_NUMERO_UNICO", 
               "OK_EXACTO", "OK_NORMALIZADO", "OK_COMPACTO", "OK_RECOMPUESTO"}
STATUS_SIN_NUMERO = "SIN_NUMERO_PRESUPUESTO"
STATUS_NO_ENCONTRADO = "NO_ENCONTRADO_EN_EXCEL"
STATUS_AMBIGUO = "AMBIGUO"
STATUS_AMBIGUO_SERIE = "AMBIGUO_SERIE"
STATUS_SERIE_NO_ENCONTRADA = "SERIE_NO_ENCONTRADA"

# Column definitions
REQUIRED_COLUMNS = [
    "Anio",
    "Presupuesto",
    "Cliente",
    "Sede_Hotel_Direccion",
    "OrigenAsignacion",
]

OPTIONAL_COLUMNS = ["Serie", "ClaveApp", "Referencia", "ClaveInterna"]

# Regex patterns
_SEPARATOR_RE = re.compile(r"[-./\\_()\[\]#]+")
_SPACES_RE = re.compile(r"\s+")
_CONTROL_RE = re.compile(r"[\x00-\x1f\x7f]")
_INVALID_PATH_CHARS_RE = re.compile(r'[<>:"/\\|?*]+')
_NUMBER_RE = re.compile(r"\d+")
_SERIE_RE = re.compile(r"\b([A-Z]{1,3})\b")


@dataclass(slots=True)
class FactusolBudgetRecord:
    anio: str
    serie: str  # Nueva: Serie del presupuesto
    presupuesto: str
    cliente: str
    sede_hotel_direccion: str
    referencia: str = ""
    origen_asignacion: str = ""
    clave_interna: str = ""
    clave_app: str = ""  # Nueva: ClaveApp = Anio-Serie-Presupuesto


@dataclass(slots=True)
class BudgetCandidate:
    raw_text: str
    normalized_text: str
    compact_text: str
    detected_budget: str
    confidence: float
    source_level: str
    reason: str
    detected_serie: str = ""  # Nueva: Serie detectada en la ruta


@dataclass(slots=True)
class MatchResult:
    status: str
    record: FactusolBudgetRecord | None = None
    candidate: BudgetCandidate | None = None
    year: str | None = None
    presupuesto_detectado: str | None = None
    serie_detectada: str = ""  # Nueva
    serie_excel: str = ""  # Nueva
    clave_app: str = ""  # Nueva
    cliente: str = ""
    sede_hotel_direccion: str = ""
    referencia: str = ""
    origen_asignacion: str = ""
    clave_interna: str = ""
    tipo_documento: str = ""
    confidence: float = 0.0
    source_level: str = ""
    reason: str = ""
    texto_detectado: str = ""
    duplicado_anio_presupuesto: str = ""

    @property
    def is_ok(self) -> bool:
        return self.status in OK_STATUSES and self.record is not None


@dataclass(slots=True)
class FactusolMappingIndex:
    records: list[FactusolBudgetRecord]
    # Key: (anio, serie, presupuesto) -> list of records (should be 1:1 usually)
    mapping_by_clave_app: dict[str, list[FactusolBudgetRecord]] = field(default_factory=dict)
    # Key: (anio, presupuesto) -> list of records (for duplicate detection)
    mapping_by_anio_presupuesto: dict[tuple[str, str], list[FactusolBudgetRecord]] = field(default_factory=dict)
    # Key: presupuesto -> list of records (for unique budget detection)
    records_by_presupuesto: dict[str, list[FactusolBudgetRecord]] = field(default_factory=dict)
    years: set[str] = field(default_factory=set)
    budgets: set[str] = field(default_factory=set)
    series: set[str] = field(default_factory=set)
    # (anio, presupuesto) tuples that have multiple series
    anio_presupuesto_with_multiple_series: set[tuple[str, str]] = field(default_factory=set)

    @classmethod
    def from_records(cls, records: Iterable[FactusolBudgetRecord]) -> "FactusolMappingIndex":
        materialized = list(records)
        by_clave_app: dict[str, list[FactusolBudgetRecord]] = defaultdict(list)
        by_anio_presupuesto: dict[tuple[str, str], list[FactusolBudgetRecord]] = defaultdict(list)
        by_presupuesto: dict[str, list[FactusolBudgetRecord]] = defaultdict(list)
        years: set[str] = set()
        budgets: set[str] = set()
        series: set[str] = set()
        anio_presupuesto_series: dict[tuple[str, str], set[str]] = defaultdict(set)

        for record in materialized:
            clave_app = record.clave_app
            key_anio_pres = (record.anio, record.presupuesto)
            
            by_clave_app[clave_app].append(record)
            by_anio_presupuesto[key_anio_pres].append(record)
            by_presupuesto[record.presupuesto].append(record)
            years.add(record.anio)
            budgets.add(record.presupuesto)
            series.add(record.serie)
            anio_presupuesto_series[key_anio_pres].add(record.serie)

        # Find (anio, presupuesto) pairs with multiple series
        multi_series = {
            k for k, v in anio_presupuesto_series.items() if len(v) > 1
        }
        
        return cls(
            records=materialized,
            mapping_by_clave_app=dict(by_clave_app),
            mapping_by_anio_presupuesto=dict(by_anio_presupuesto),
            records_by_presupuesto=dict(by_presupuesto),
            years=years,
            budgets=budgets,
            series=series,
            anio_presupuesto_with_multiple_series=multi_series,
        )

    def is_unique_budget_in_year(self, anio: str, presupuesto: str) -> bool:
        """Check if presupuesto is unique within the year (no multiple series)."""
        key = (anio, presupuesto)
        return key not in self.anio_presupuesto_with_multiple_series

    def get_records_by_clave_app(self, clave_app: str) -> list[FactusolBudgetRecord]:
        return self.mapping_by_clave_app.get(clave_app, [])

    def get_records_by_anio_serie_presupuesto(self, anio: str, serie: str, presupuesto: str) -> list[FactusolBudgetRecord]:
        clave_app = f"{anio}-{serie}-{presupuesto}"
        return self.get_records_by_clave_app(clave_app)


class FactusolMappingLoader:
    """Loads the simplified FactuSOL Excel mapping."""

    sheet_name = "Mapping_FactuSOL"
    DEFAULT_SERIE = "GENERAL"

    def load_mapping(self, excel_path: Path) -> FactusolMappingIndex:
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise RuntimeError("openpyxl no esta instalado. pip install openpyxl") from exc

        if not excel_path.exists():
            raise FileNotFoundError(f"Excel de mapping no encontrado: {excel_path}")

        workbook = load_workbook(excel_path, read_only=True, data_only=True)
        try:
            if self.sheet_name not in workbook.sheetnames:
                raise ValueError(f"El Excel debe contener la hoja '{self.sheet_name}'.")

            sheet = workbook[self.sheet_name]
            rows = sheet.iter_rows(values_only=True)
            try:
                header_row = next(rows)
            except StopIteration as exc:
                raise ValueError("La hoja Mapping_FactuSOL esta vacia.") from exc

            headers = [_cell_to_text(value) for value in header_row]
            header_positions = {name: idx for idx, name in enumerate(headers) if name}
            missing = [column for column in REQUIRED_COLUMNS if column not in header_positions]
            if missing:
                raise ValueError(
                    "Faltan columnas obligatorias en Mapping_FactuSOL: "
                    + ", ".join(missing)
                )

            records: list[FactusolBudgetRecord] = []
            for row in rows:
                data = {
                    column: _cell_to_text(row[header_positions[column]])
                    if header_positions[column] < len(row)
                    else ""
                    for column in REQUIRED_COLUMNS
                }
                for column in OPTIONAL_COLUMNS:
                    idx = header_positions.get(column)
                    data[column] = _cell_to_text(row[idx]) if idx is not None and idx < len(row) else ""

                if not any(data.values()):
                    continue

                anio = data["Anio"].strip()
                serie = data["Serie"].strip() or self.DEFAULT_SERIE
                presupuesto = data["Presupuesto"].strip()
                
                # Build or use ClaveApp
                clave_app = data.get("ClaveApp", "").strip()
                if not clave_app:
                    clave_app = f"{anio}-{serie}-{presupuesto}"
                
                cliente = data["Cliente"].strip()
                sede = data["Sede_Hotel_Direccion"].strip() or cliente
                
                records.append(
                    FactusolBudgetRecord(
                        anio=anio,
                        serie=serie,
                        presupuesto=presupuesto,
                        cliente=cliente,
                        sede_hotel_direccion=sede,
                        referencia=data.get("Referencia", "").strip(),
                        origen_asignacion=data.get("OrigenAsignacion", "").strip(),
                        clave_interna=data.get("ClaveInterna", "").strip(),
                        clave_app=clave_app,
                    )
                )
        finally:
            workbook.close()

        return FactusolMappingIndex.from_records(records)


def load_mapping(excel_path: Path) -> FactusolMappingIndex:
    return FactusolMappingLoader().load_mapping(excel_path)


def normalize_text(value: str) -> str:
    raw = "" if value is None else str(value)
    without_accents = _strip_accents(raw).upper()
    separated = _SEPARATOR_RE.sub(" ", without_accents)
    separated = _CONTROL_RE.sub(" ", separated)
    return _SPACES_RE.sub(" ", separated).strip()


def compact_text(value: str) -> str:
    return normalize_text(value).replace(" ", "")


def sanitize_path_part(value: str, max_len: int = 80) -> str:
    text = _strip_accents("" if value is None else str(value))
    text = _CONTROL_RE.sub(" ", text)
    text = text.replace("\r", " ").replace("\n", " ").replace("\t", " ")
    text = _INVALID_PATH_CHARS_RE.sub(" ", text)
    text = _SPACES_RE.sub(" ", text).strip(" .")
    if max_len > 0:
        text = text[:max_len].rstrip(" .")
    return text or "_SIN_NOMBRE"


def detect_year_from_path(path: Path, allowed_years: set[str]) -> str | None:
    if not allowed_years:
        return None
    for part in path.parts:
        normalized = normalize_text(part)
        if normalized in allowed_years:
            return normalized
    for number in _NUMBER_RE.findall(normalize_text(str(path))):
        if number in allowed_years:
            return number
    return None


def detect_serie_from_path(path: Path, available_series: set[str]) -> str | None:
    """Detect Serie from path. Returns None if not found or ambiguous."""
    if not available_series:
        return None
    
    path_text = normalize_text(str(path))
    
    # Look for serie patterns like "A-", "_B", " SERIE A ", etc.
    for serie in available_series:
        if serie == "GENERAL":
            continue
        
        # Patterns that indicate a serie in the path
        serie_patterns = [
            f"-{serie}-",          # e.g., "A-250076"
            f"_{serie}_",          # e.g., "A_250076"
            f" {serie} ",          # e.g., "SERIE A 250076"
            f"{serie}-",           # e.g., "A-250076" at start
            f"-{serie}$",          # e.g., "250076-A" at end
            f"^{serie}-",          # e.g., "A-250076" at start
        ]
        
        for pattern in serie_patterns:
            if pattern.lower() in path_text.lower():
                return serie
    
    return None


def _extract_numbers_from_text(text: str) -> list[str]:
    """Extract potential budget numbers (5+ digits) from text."""
    return [n for n in _NUMBER_RE.findall(text) if len(n) >= 5]


def detect_budget_with_serie(path: Path, mapping_index: FactusolMappingIndex) -> list[BudgetCandidate]:
    """Detect budget candidates including serie detection - optimized version."""
    candidates: list[BudgetCandidate] = []
    seen: set[str] = set()

    for raw_text, source_level, source_rank in _path_source_texts(path):
        if not raw_text:
            continue
        normalized = normalize_text(raw_text)
        compact = compact_text(raw_text)

        # Extract numbers from this text - these are potential budget numbers
        numbers = _extract_numbers_from_text(normalized)
        if not numbers:
            continue

        # For each number found, check if it's a known budget
        for number in numbers:
            if number in mapping_index.budgets and number not in seen:
                seen.add(number)
                reason = _match_reason_for_budget(raw_text, number)
                if reason:
                    # Try to detect serie from the same text
                    detected_serie = detect_serie_from_path(Path(raw_text), mapping_index.series)

                    candidates.append(
                        BudgetCandidate(
                            raw_text=raw_text,
                            normalized_text=normalized,
                            compact_text=compact,
                            detected_budget=number,
                            detected_serie=detected_serie or "",
                            confidence=_confidence_for(reason, source_rank),
                            source_level=source_level,
                            reason=reason,
                        )
                    )

    candidates.sort(key=lambda item: item.confidence, reverse=True)
    return candidates


# Keep backward compatibility
def detect_budget_candidates(path: Path, mapping_index: FactusolMappingIndex) -> list[BudgetCandidate]:
    return detect_budget_with_serie(path, mapping_index)


def resolve_budget_match(
    path: Path,
    mapping_index: FactusolMappingIndex,
    allowed_years: set[str] | None = None,
) -> MatchResult:
    effective_years = {str(year).strip() for year in allowed_years or mapping_index.years if str(year).strip()}
    detected_year = detect_year_from_path(path, effective_years)
    candidates = detect_budget_candidates(path, mapping_index)

    if not candidates:
        unknown = _unknown_budget_numbers(path, mapping_index)
        if unknown:
            return MatchResult(
                status=STATUS_NO_ENCONTRADO,
                year=detected_year,
                presupuesto_detectado=unknown[0],
                reason=STATUS_NO_ENCONTRADO,
                texto_detectado=unknown[0],
            )
        return MatchResult(
            status=STATUS_SIN_NUMERO,
            year=detected_year,
            reason=STATUS_SIN_NUMERO,
        )

    # Get the best candidate
    best = candidates[0]
    presupuesto = best.detected_budget
    
    # Step 1: Check if we can detect a serie from path
    detected_serie = best.detected_serie
    
    # Step 2: Get all records matching (year, presupuesto)
    anio_records = mapping_index.mapping_by_anio_presupuesto.get((detected_year or "", presupuesto), [])
    if not anio_records and detected_year:
        # Try with any year
        for year in mapping_index.years:
            anio_records = mapping_index.mapping_by_anio_presupuesto.get((year, presupuesto), [])
            if anio_records:
                detected_year = year
                break
    
    if not anio_records:
        return MatchResult(
            status=STATUS_NO_ENCONTRADO,
            candidate=best,
            year=detected_year,
            presupuesto_detectado=presupuesto,
            confidence=best.confidence,
            source_level=best.source_level,
            reason=STATUS_NO_ENCONTRADO,
            texto_detectado=best.raw_text,
        )
    
    # Step 3: Check uniqueness
    has_multiple_series = (detected_year or "", presupuesto) in mapping_index.anio_presupuesto_with_multiple_series
    
    if not has_multiple_series:
        # Only one serie for this presupuesto+year - use it directly
        record = anio_records[0]
        return _ok_result(
            status="OK_NUMERO_UNICO",
            record=record,
            candidate=best,
            mapping_index=mapping_index,
            confidence=best.confidence,
            reason="Presupuesto unico para el anio, serie automatica.",
        )
    
    # Step 4: Multiple series exist - need to detect serie from path
    if detected_serie:
        # Try exact match with detected serie
        matching_records = [r for r in anio_records if r.serie == detected_serie]
        if matching_records:
            record = matching_records[0]
            return _ok_result(
                status="OK_ANIO_SERIE_PRESUPUESTO",
                record=record,
                candidate=best,
                mapping_index=mapping_index,
                confidence=best.confidence,
                reason=f"Serie {detected_serie} detectada en ruta.",
            )
        else:
            # Serie detected but doesn't match any record
            return MatchResult(
                status=STATUS_SERIE_NO_ENCONTRADA,
                candidate=best,
                year=detected_year,
                presupuesto_detectado=presupuesto,
                serie_detectada=detected_serie,
                confidence=best.confidence,
                source_level=best.source_level,
                reason=f"Serie {detected_serie} detectada pero no existe en Excel para este presupuesto.",
                texto_detectado=best.raw_text,
            )
    
    # Step 5: No serie detected and multiple exist - AMBIGUO
    return MatchResult(
        status=STATUS_AMBIGUO_SERIE,
        candidate=best,
        year=detected_year,
        presupuesto_detectado=presupuesto,
        confidence=best.confidence,
        source_level=best.source_level,
        reason=f"Multiple series ({len(anio_records)}) para presupuesto {presupuesto} en anio {detected_year}. Serie no detectable en ruta.",
        texto_detectado=best.raw_text,
    )


def categorize_document_type(path: Path) -> str:
    ext = path.suffix.lower()
    if ext == ".pdf":
        return "PDF"
    if ext in {".xls", ".xlsx", ".xlsm", ".csv"}:
        return "EXCEL"
    if ext in {".jpg", ".jpeg", ".png", ".tif", ".tiff", ".bmp", ".gif"}:
        return "IMAGENES"
    if ext in {".msg", ".eml", ".pst", ".ost"}:
        return "CORREOS"
    if ext in {".dwg", ".dxf", ".skp", ".rvt", ".ifc", ".pln", ".3dm"}:
        return "PLANOS"
    if ext in {".doc", ".docx"}:
        return "WORD"
    if ext in {".zip", ".rar", ".7z"}:
        return "ZIP"
    return "OTROS"


def build_factusol_client_budget_destination_path(
    src: Path,
    dest_root: Path,
    source_root: Path,
    match_result: MatchResult,
    unmatched_dir: str = "_REVISION",
) -> Path:
    if match_result.is_ok and match_result.record is not None:
        record = match_result.record
        sede = record.sede_hotel_direccion or record.cliente
        return (
            dest_root
            / sanitize_path_part(record.anio)
            / sanitize_path_part(record.cliente)
            / sanitize_path_part(sede)
            / sanitize_path_part(f"Presupuesto {record.presupuesto}")
            / categorize_document_type(src)
            / src.name
        )

    review_root = dest_root / sanitize_path_part(unmatched_dir)
    relative_parent = _relative_parent(src, source_root)

    if match_result.status == STATUS_SIN_NUMERO:
        return _append_relative(review_root / "_SIN_NUMERO_PRESUPUESTO", relative_parent) / src.name
    if match_result.status == STATUS_NO_ENCONTRADO:
        year = sanitize_path_part(match_result.year or "DESCONOCIDO")
        return _append_relative(review_root / "_NO_ENCONTRADO_EN_EXCEL" / year, relative_parent) / src.name
    if match_result.status == STATUS_AMBIGUO_SERIE:
        return _append_relative(review_root / "_AMBIGUO_SERIE", relative_parent) / src.name
    if match_result.status == STATUS_SERIE_NO_ENCONTRADA:
        return _append_relative(review_root / "_SERIE_NO_ENCONTRADA", relative_parent) / src.name
    return _append_relative(review_root / "_AMBIGUO", relative_parent) / src.name


def _cell_to_text(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _strip_accents(value: str) -> str:
    return "".join(
        char
        for char in unicodedata.normalize("NFKD", value)
        if not unicodedata.combining(char)
    )


def _path_source_texts(path: Path) -> list[tuple[str, str, int]]:
    items: list[tuple[str, str, int]] = []
    parent_name = path.parent.name
    if parent_name:
        items.append((parent_name, "parent_folder", 0))
    if path.stem:
        items.append((path.stem, "filename", 1))
    for parent in path.parents[1:5]:
        if parent.name:
            items.append((parent.name, "ancestor_folder", 2))
    tail_parts = path.parts[-7:] if len(path.parts) > 7 else path.parts
    if tail_parts:
        items.append((" ".join(tail_parts), "relative_path", 3))
    items.append((str(path), "full_path", 4))
    return items


def _match_reason_for_budget(raw_text: str, budget: str) -> str | None:
    normalized = normalize_text(raw_text)
    compact = normalized.replace(" ", "")
    budget_norm = normalize_text(budget)
    budget_compact = budget_norm.replace(" ", "")

    if normalized == budget_norm:
        return "OK_EXACTO"
    if _is_recomposed_match(normalized, budget_compact):
        return "OK_RECOMPUESTO"
    if budget_compact and budget_compact in compact and compact != budget_compact:
        return "OK_COMPACTO"
    if budget_norm in normalized.split():
        return "OK_NORMALIZADO"
    if compact == budget_compact and normalized != budget_norm:
        return "OK_RECOMPUESTO"
    return None


def _is_recomposed_match(normalized_text: str, budget_compact: str) -> bool:
    numbers = _NUMBER_RE.findall(normalized_text)
    if len(numbers) < 2:
        return False
    for start in range(len(numbers)):
        combined = ""
        for number in numbers[start:]:
            combined += number
            if combined == budget_compact:
                return True
            if len(combined) >= len(budget_compact):
                break
    return False


def _confidence_for(reason: str, source_rank: int) -> float:
    base = {
        "OK_EXACTO": 0.99,
        "OK_NORMALIZADO": 0.94,
        "OK_COMPACTO": 0.90,
        "OK_RECOMPUESTO": 0.88,
    }.get(reason, 0.85)
    return max(0.10, base - (source_rank * 0.05))


def _unknown_budget_numbers(path: Path, mapping_index: FactusolMappingIndex) -> list[str]:
    found: list[str] = []
    seen: set[str] = set()
    for raw_text, _, _ in _path_source_texts(path):
        normalized = normalize_text(raw_text)
        numbers = _NUMBER_RE.findall(normalized)
        for number in numbers:
            if len(number) >= 5 and number not in mapping_index.budgets and number not in mapping_index.years:
                if number not in seen:
                    seen.add(number)
                    found.append(number)
        for recomposed in _recomposed_numbers(numbers):
            if recomposed not in mapping_index.budgets and recomposed not in mapping_index.years:
                if recomposed not in seen:
                    seen.add(recomposed)
                    found.append(recomposed)
    return found


def _recomposed_numbers(numbers: Sequence[str]) -> Iterable[str]:
    for start in range(len(numbers)):
        combined = ""
        for number in numbers[start:]:
            combined += number
            if len(combined) >= 5:
                yield combined
            if len(combined) >= 8:
                break


def _ok_result(
    status: str,
    record: FactusolBudgetRecord,
    candidate: BudgetCandidate,
    mapping_index: FactusolMappingIndex,
    confidence: float,
    reason: str | None = None,
) -> MatchResult:
    return MatchResult(
        status=status,
        record=record,
        candidate=candidate,
        year=record.anio,
        presupuesto_detectado=record.presupuesto,
        serie_detectada=candidate.detected_serie,
        serie_excel=record.serie,
        clave_app=record.clave_app,
        cliente=record.cliente,
        sede_hotel_direccion=record.sede_hotel_direccion or record.cliente,
        referencia=record.referencia,
        origen_asignacion=record.origen_asignacion,
        clave_interna=record.clave_interna,
        confidence=confidence,
        source_level=candidate.source_level,
        reason=reason or candidate.reason,
        texto_detectado=candidate.raw_text,
    )


def _relative_parent(src: Path, source_root: Path) -> tuple[str, ...]:
    try:
        relative = src.relative_to(source_root)
    except ValueError:
        relative = Path(src.name)
    return tuple(sanitize_path_part(part) for part in relative.parts[:-1])


def _append_relative(base: Path, parts: Sequence[str]) -> Path:
    result = base
    for part in parts:
        result = result / part
    return result
