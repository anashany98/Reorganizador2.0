"""Generador de Excel de auditoria post-escaneo."""

from __future__ import annotations

import logging
from datetime import datetime
from pathlib import Path
from typing import Optional

LOGGER = logging.getLogger(__name__)

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    Workbook = None  # type: ignore

_HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
_HEADER_FILL = PatternFill(start_color="1A2332", end_color="1A2332", fill_type="solid")
_HEADER_ALIGN = Alignment(horizontal="left", vertical="center", wrap_text=True)

_TITLE_FONT = Font(name="Calibri", bold=True, size=14, color="14E0A8")
_SUBTITLE_FONT = Font(name="Calibri", size=11, color="8A98A8")
_DATA_FONT = Font(name="Calibri", size=10)
_MONO_FONT = Font(name="Consolas", size=10)

_THIN_BORDER = Border(bottom=Side(style="thin", color="2A3747"))

_ERR_FILL = PatternFill(start_color="3D1218", end_color="3D1218", fill_type="solid")
_ERR_FONT = Font(name="Calibri", size=10, color="F44B67")
_ALT_FILL = PatternFill(start_color="0E131B", end_color="0E131B", fill_type="solid")

_COLUMNS = [
    ("Archivo", 28), ("Extension", 10), ("Tipo MIME", 22),
    ("Tamano (bytes)", 16), ("Creado", 18), ("Modificado", 18),
    ("Hash", 18), ("Hash valor", 66), ("Ruta origen", 55),
    ("Ruta destino", 55), ("Gestor", 18), ("Proyecto", 16),
    ("Anio", 10), ("Presupuesto detectado", 18), ("Cliente", 28),
    ("Sede_Hotel_Direccion", 32), ("Referencia", 36),
    ("OrigenAsignacion", 18), ("ClaveInterna", 22), ("TipoDocumento", 16),
    ("MatchStatus", 20), ("Confianza", 12), ("MatchSource", 16),
    ("MatchReason", 40), ("TextoDetectado", 28), ("Duplicado", 12),
    ("Accion", 10), ("Estado", 10), ("Error", 40), ("Verificado", 12),
]


def _style_header(ws, num_cols):
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=1, column=col)
        if _HEADER_FONT: cell.font = _HEADER_FONT
        if _HEADER_FILL: cell.fill = _HEADER_FILL
        cell.alignment = _HEADER_ALIGN
        cell.border = _THIN_BORDER


def _auto_width(ws, widths):
    for idx, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = max(w, 10)


def _write_data_rows(ws, rows, start_row=2):
    keys = [
        "file_name", "extension", "mime_type", "size_bytes",
        "created_time", "modified_time", "hash_algo", "hash_value",
        "src_path", "dst_path", "gestor", "proyecto",
        "year", "presupuesto_detectado", "cliente", "sede_hotel_direccion",
        "referencia", "origen_asignacion", "clave_interna", "tipo_documento",
        "match_status", "match_confidence", "match_source", "match_reason",
        "texto_detectado", "duplicado_anio_presupuesto",
        "action", "action_status", "error", "verified",
    ]
    for r_idx, row in enumerate(rows, start=start_row):
        is_alt = (r_idx % 2 == 0)
        is_err = row.get("action_status") == "error"
        for c_idx, key in enumerate(keys, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=row.get(key, ""))
            if _ERR_FONT and is_err: cell.font = _ERR_FONT
            else: cell.font = _DATA_FONT
            cell.border = _THIN_BORDER
            if is_err and _ERR_FILL: cell.fill = _ERR_FILL
            elif is_alt and _ALT_FILL: cell.fill = _ALT_FILL


def generate_audit_excel(db_path, output_path, source_label="", dest_label=""):
    if Workbook is None:
        raise RuntimeError("openpyxl no esta instalado. pip install openpyxl")

    import sqlite3

    if not db_path.exists():
        raise FileNotFoundError(f"Base de datos no encontrada: {db_path}")

    conn = sqlite3.connect(str(db_path))
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()

    cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='files'")
    if not cursor.fetchone():
        conn.close()
        raise RuntimeError("La base de datos no contiene la tabla 'files'.")

    cursor.execute("SELECT * FROM files ORDER BY id")
    all_rows = [dict(row) for row in cursor.fetchall()]
    error_rows = [r for r in all_rows if r.get("action_status") == "error"]

    processed = len(all_rows)
    errors = len(error_rows)
    skipped = sum(1 for r in all_rows if r.get("action") == "skip")
    copied = sum(1 for r in all_rows if r.get("action") == "copy")
    moved = sum(1 for r in all_rows if r.get("action") == "move")
    ok_matches = sum(1 for r in all_rows if str(r.get("match_status") or "").startswith("OK_"))
    without_number = sum(1 for r in all_rows if r.get("match_status") == "SIN_NUMERO_PRESUPUESTO")
    not_found = sum(1 for r in all_rows if r.get("match_status") == "NO_ENCONTRADO_EN_EXCEL")
    ambiguous = sum(1 for r in all_rows if r.get("match_status") == "AMBIGUO")

    wb = Workbook()

    # ---- Hoja 1: Resumen ----
    ws1 = wb.active
    ws1.title = "Resumen"
    ws1.merge_cells("A1:B1")
    ws1["A1"].value = "Informe de Auditoria - Reorganizador 2.0"
    ws1["A1"].font = _TITLE_FONT
    ws1.merge_cells("A2:B2")
    ws1["A2"].value = f"Generado: {datetime.now():%Y-%m-%d %H:%M:%S}"
    ws1["A2"].font = _SUBTITLE_FONT

    summary = [
        ("Origen", source_label or "-"),
        ("Destino", dest_label or "-"),
        ("", ""),
        ("Total archivos", processed),
        ("Copiados", copied),
        ("Movidos", moved),
        ("Omitidos (skip)", skipped),
        ("Errores", errors),
        ("Match OK", ok_matches),
        ("Sin numero presupuesto", without_number),
        ("No encontrado en Excel", not_found),
        ("Ambiguos", ambiguous),
    ]
    for r_idx, (label, value) in enumerate(summary, start=4):
        c1 = ws1.cell(row=r_idx, column=1, value=label)
        c2 = ws1.cell(row=r_idx, column=2, value=value)
        c1.font = Font(name="Calibri", bold=True, size=11, color="8A98A8")
        c2.font = Font(name="Calibri", size=11, color="E4E8EE")
        if label == "Errores" and errors > 0:
            c2.font = _ERR_FONT
    ws1.column_dimensions["A"].width = 22
    ws1.column_dimensions["B"].width = 28

    # ---- Hoja 2: Detalle ----
    ws2 = wb.create_sheet("Detalle")
    for c_idx, (name, _) in enumerate(_COLUMNS, start=1):
        ws2.cell(row=1, column=c_idx, value=name)
    _style_header(ws2, len(_COLUMNS))
    _write_data_rows(ws2, all_rows)
    _auto_width(ws2, [w for _, w in _COLUMNS])
    ws2.auto_filter.ref = ws2.dimensions
    ws2.freeze_panes = "A2"

    # ---- Hoja 3: Errores ----
    if error_rows:
        ws3 = wb.create_sheet("Errores")
        for c_idx, (name, _) in enumerate(_COLUMNS, start=1):
            ws3.cell(row=1, column=c_idx, value=name)
        _style_header(ws3, len(_COLUMNS))
        _write_data_rows(ws3, error_rows)
        _auto_width(ws3, [w for _, w in _COLUMNS])
        ws3.auto_filter.ref = ws3.dimensions
        ws3.freeze_panes = "A2"

    # ---- Hoja 4: Por Gestor ----
    cursor.execute("""
        SELECT COALESCE(gestor, 'Sin gestor') AS gestor,
               COUNT(*) AS total,
               SUM(CASE WHEN action = 'copy' THEN 1 ELSE 0 END) AS copiados,
               SUM(CASE WHEN action = 'move' THEN 1 ELSE 0 END) AS movidos,
               SUM(CASE WHEN action = 'skip' THEN 1 ELSE 0 END) AS omitidos,
               SUM(CASE WHEN action_status = 'error' THEN 1 ELSE 0 END) AS errores
        FROM files GROUP BY gestor ORDER BY total DESC
    """)
    gestor_rows = [dict(r) for r in cursor.fetchall()]

    ws4 = wb.create_sheet("Por Gestor")
    gh = ["Gestor", "Total", "Copiados", "Movidos", "Omitidos", "Errores"]
    for c_idx, h in enumerate(gh, start=1):
        ws4.cell(row=1, column=c_idx, value=h)
    _style_header(ws4, len(gh))
    for r_idx, row in enumerate(gestor_rows, start=2):
        for c_idx, key in enumerate(["gestor", "total", "copiados", "movidos", "omitidos", "errores"], start=1):
            cell = ws4.cell(row=r_idx, column=c_idx, value=row.get(key, 0))
            cell.font = _DATA_FONT; cell.border = _THIN_BORDER
            if r_idx % 2 == 0: cell.fill = _ALT_FILL
            if key == "errores" and row.get(key, 0) > 0: cell.font = _ERR_FONT
    _auto_width(ws4, [22, 12, 12, 12, 12, 12])
    ws4.auto_filter.ref = ws4.dimensions

    # ---- Hoja 5: Por Proyecto ----
    cursor.execute("""
        SELECT COALESCE(gestor, 'Sin gestor') AS gestor,
               COALESCE(proyecto, 'Sin proyecto') AS proyecto,
               COUNT(*) AS total,
               SUM(CASE WHEN action_status = 'error' THEN 1 ELSE 0 END) AS errores
        FROM files GROUP BY gestor, proyecto ORDER BY total DESC
    """)
    proyecto_rows = [dict(r) for r in cursor.fetchall()]

    ws5 = wb.create_sheet("Por Proyecto")
    ph = ["Gestor", "Proyecto", "Total", "Errores"]
    for c_idx, h in enumerate(ph, start=1):
        ws5.cell(row=1, column=c_idx, value=h)
    _style_header(ws5, len(ph))
    for r_idx, row in enumerate(proyecto_rows, start=2):
        for c_idx, key in enumerate(["gestor", "proyecto", "total", "errores"], start=1):
            cell = ws5.cell(row=r_idx, column=c_idx, value=row.get(key, 0))
            cell.font = _DATA_FONT; cell.border = _THIN_BORDER
            if r_idx % 2 == 0: cell.fill = _ALT_FILL
            if key == "errores" and row.get(key, 0) > 0: cell.font = _ERR_FONT
    _auto_width(ws5, [22, 22, 12, 12])
    ws5.auto_filter.ref = ws5.dimensions

    conn.close()
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output_path))
    LOGGER.info("Excel de auditoria generado: %s", output_path)
    return output_path
